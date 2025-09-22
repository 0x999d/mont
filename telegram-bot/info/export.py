from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from datetime import datetime

from io import BytesIO

from typing import List, Dict, Any
from utils import sync_to_async


@sync_to_async
def export_history(data: List[Dict[str, Any]]) -> bytes:
    if not data:
        raise ValueError("Empty data provided.")

    data = sorted(data, key=lambda x: x['date'])

    wb = Workbook()
    ws = wb.active
    ws.title = "Availability Report"

    headers = ["date", "latency", "ok"]
    ws.append(headers)

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    font = Font(bold=True)

    for row in data:
        timestamp = row['date']
        dt = datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")
        latency = row['latency'] if row['latency'] is not None else "unknown"
        status = "up" if row['is_ok'] else "down"

        excel_row = [dt, latency, status]
        ws.append(excel_row)

        fill = green if row['is_ok'] else red
        for col in range(1, 4):
            ws.cell(row=ws.max_row, column=col).fill = fill

    for coll, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=coll)
        cell.font = font

        max_length = max(len(str(cell.value)) for cell in ws[
            ws.cell(row=1, column=coll).column_letter
        ])
        ws.column_dimensions[get_column_letter(coll)].width = max(
            max_length + 2, 15
        )

    buff = BytesIO()
    wb.save(buff)
    return buff.getvalue()